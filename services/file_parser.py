"""文件解析服务：支持 CSV、TXT 和 Excel 格式"""
import csv
import io
import os
import chardet
from openpyxl import load_workbook


def detect_encoding(filepath, sample_size=32768):
    """检测文件编码"""
    with open(filepath, 'rb') as f:
        raw_data = f.read(sample_size)
        result = chardet.detect(raw_data)
        encoding = result['encoding']
        confidence = result['confidence']
        # 如果置信度不高或检测为 ascii，尝试常见中文编码
        if confidence < 0.7 or encoding.lower() == 'ascii':
            for enc in ['utf-8-sig', 'gbk', 'gb2312', 'gb18030']:
                try:
                    raw_data.decode(enc)
                    return enc
                except UnicodeDecodeError:
                    continue
        return encoding or 'utf-8'


def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions


def parse_csv(filepath, corpus_type='question'):
    """
    解析 CSV 文件
    corpus_type='question': 只需要 question 列
    corpus_type='qa': 需要 question 和 answer 列
    """
    items = []
    encoding = detect_encoding(filepath)
    with open(filepath, 'r', encoding=encoding) as f:
        # 尝试检测分隔符
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
        except csv.Error:
            dialect = csv.excel

        reader = csv.DictReader(f, dialect=dialect)
        fieldnames = [name.strip().lower() for name in (reader.fieldnames or [])]

        # 找到 question 列
        q_col = None
        for name in fieldnames:
            if name in ('question', 'q', '问题', 'query', 'prompt', 'input'):
                q_col = reader.fieldnames[fieldnames.index(name)]
                break

        # 找到 answer 列
        a_col = None
        if corpus_type == 'qa':
            for name in fieldnames:
                if name in ('answer', 'a', '答案', 'response', 'output', 'reply'):
                    a_col = reader.fieldnames[fieldnames.index(name)]
                    break

        if q_col is None:
            # 如果没有列头匹配，尝试用第一列作为 question
            f.seek(0)
            reader = csv.reader(f, dialect=dialect)
            rows = list(reader)
            if len(rows) > 0:
                has_header = not _looks_like_data(rows[0][0]) if rows[0] else False
                start = 1 if has_header else 0
                for row in rows[start:]:
                    if not row or not row[0].strip():
                        continue
                    item = {'question': row[0].strip()}
                    if corpus_type == 'qa' and len(row) > 1:
                        item['answer'] = row[1].strip()
                    items.append(item)
            return items

        for row in reader:
            q = row.get(q_col, '').strip()
            if not q:
                continue
            item = {'question': q}
            if corpus_type == 'qa' and a_col:
                item['answer'] = row.get(a_col, '').strip()
            items.append(item)

    return items


def parse_txt(filepath, corpus_type='question'):
    """
    解析 TXT 文件
    每行一个 question；如果是 QA 对，用 \t 或 ||| 分隔
    """
    items = []
    encoding = detect_encoding(filepath)
    with open(filepath, 'r', encoding=encoding) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            if corpus_type == 'qa':
                # 尝试不同分隔符
                parts = None
                for sep in ['\t', '|||', '|', '::']:
                    if sep in line:
                        parts = line.split(sep, 1)
                        break
                if parts and len(parts) == 2:
                    items.append({
                        'question': parts[0].strip(),
                        'answer': parts[1].strip()
                    })
                else:
                    items.append({'question': line})
            else:
                items.append({'question': line})

    return items


def parse_excel(filepath, corpus_type='question'):
    """
    解析 Excel 文件（.xlsx / .xls）
    corpus_type='question': 只需要 question 列
    corpus_type='qa': 需要 question 和 answer 列
    """
    items = []
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return items

    # 检测表头行，找 question / answer 列索引
    header = [str(cell).strip().lower() if cell is not None else '' for cell in rows[0]]

    q_idx = None
    for i, name in enumerate(header):
        if name in ('question', 'q', '问题', 'query', 'prompt', 'input'):
            q_idx = i
            break

    a_idx = None
    if corpus_type == 'qa':
        for i, name in enumerate(header):
            if name in ('answer', 'a', '答案', 'response', 'output', 'reply'):
                a_idx = i
                break

    if q_idx is not None:
        # 有匹配的表头，从第二行开始读数据
        for row in rows[1:]:
            q = str(row[q_idx]).strip() if q_idx < len(row) and row[q_idx] is not None else ''
            if not q:
                continue
            item = {'question': q}
            if corpus_type == 'qa' and a_idx is not None and a_idx < len(row) and row[a_idx] is not None:
                item['answer'] = str(row[a_idx]).strip()
            items.append(item)
    else:
        # 没有匹配的表头，判断第一行是否为表头
        has_header = not _looks_like_data(str(rows[0][0])) if rows[0][0] is not None else False
        start = 1 if has_header else 0
        for row in rows[start:]:
            if not row or row[0] is None or not str(row[0]).strip():
                continue
            item = {'question': str(row[0]).strip()}
            if corpus_type == 'qa' and len(row) > 1 and row[1] is not None:
                item['answer'] = str(row[1]).strip()
            items.append(item)

    return items


def parse_file(filepath, corpus_type='question'):
    """根据文件扩展名选择解析器"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.csv':
        return parse_csv(filepath, corpus_type)
    elif ext == '.txt':
        return parse_txt(filepath, corpus_type)
    elif ext in ('.xlsx', '.xls'):
        return parse_excel(filepath, corpus_type)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")


def _looks_like_data(value):
    """简单判断是否像数据行（而非表头）"""
    if len(value) > 50:
        return True
    if value.replace(' ', '').isdigit():
        return True
    return False


# ==================== 相似度对比文件解析 ====================

# 列名识别关键词
_Q_KEYWORDS = ('question', 'q', '问题', 'query', 'prompt', 'input')
_A1_KEYWORDS = ('answer1', 'a1', '答案1', 'response1', 'output1', 'reply1', 'answer_1')
_A2_KEYWORDS = ('answer2', 'a2', '答案2', 'response2', 'output2', 'reply2', 'answer_2')
_A_GENERIC_KEYWORDS = ('answer', 'a', '答案', 'response', 'output', 'reply')


def _find_similarity_columns(header):
    """从表头中识别 question / answer1 / answer2 列索引"""
    q_idx, a1_idx, a2_idx = None, None, None

    for i, name in enumerate(header):
        if q_idx is None and name in _Q_KEYWORDS:
            q_idx = i
        if a1_idx is None and name in _A1_KEYWORDS:
            a1_idx = i
        if a2_idx is None and name in _A2_KEYWORDS:
            a2_idx = i

    # 回退：如果 answer1/answer2 没有精确匹配，尝试找两个通用 answer 列
    if a1_idx is None or a2_idx is None:
        generic_matches = [i for i, name in enumerate(header)
                           if name in _A_GENERIC_KEYWORDS and i != q_idx]
        if len(generic_matches) >= 2:
            a1_idx = generic_matches[0]
            a2_idx = generic_matches[1]

    return q_idx, a1_idx, a2_idx


def parse_csv_similarity(filepath):
    """解析 CSV 文件，提取 question / answer1 / answer2 三列"""
    items = []
    encoding = detect_encoding(filepath)
    with open(filepath, 'r', encoding=encoding) as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(f, dialect=dialect)
        rows = list(reader)

    if not rows:
        return items

    header = [name.strip().lower() for name in rows[0]]
    q_idx, a1_idx, a2_idx = _find_similarity_columns(header)

    # 最终回退：位置推断（第1/2/3列）
    if q_idx is not None and a1_idx is not None and a2_idx is not None:
        start = 1
    elif len(rows[0]) >= 3:
        q_idx, a1_idx, a2_idx = 0, 1, 2
        has_header = not _looks_like_data(rows[0][0]) if rows[0] else False
        start = 1 if has_header else 0
    else:
        return items

    for row in rows[start:]:
        if len(row) <= max(q_idx, a1_idx, a2_idx):
            continue
        q = row[q_idx].strip() if row[q_idx] else ''
        a1 = row[a1_idx].strip() if row[a1_idx] else ''
        a2 = row[a2_idx].strip() if row[a2_idx] else ''
        if not q and not a1 and not a2:
            continue
        items.append({'question': q, 'answer1': a1, 'answer2': a2})

    return items


def parse_excel_similarity(filepath):
    """解析 Excel 文件，提取 question / answer1 / answer2 三列"""
    items = []
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return items

    header = [str(cell).strip().lower() if cell is not None else '' for cell in rows[0]]
    q_idx, a1_idx, a2_idx = _find_similarity_columns(header)

    if q_idx is not None and a1_idx is not None and a2_idx is not None:
        start = 1
    elif len(rows[0]) >= 3:
        q_idx, a1_idx, a2_idx = 0, 1, 2
        has_header = not _looks_like_data(str(rows[0][0])) if rows[0][0] is not None else False
        start = 1 if has_header else 0
    else:
        return items

    for row in rows[start:]:
        max_idx = max(q_idx, a1_idx, a2_idx)
        if len(row) <= max_idx:
            continue
        q = str(row[q_idx]).strip() if row[q_idx] is not None else ''
        a1 = str(row[a1_idx]).strip() if row[a1_idx] is not None else ''
        a2 = str(row[a2_idx]).strip() if row[a2_idx] is not None else ''
        if not q and not a1 and not a2:
            continue
        items.append({'question': q, 'answer1': a1, 'answer2': a2})

    return items


def parse_txt_similarity(filepath):
    """解析 TXT 文件，每行用 TAB 或 ||| 分隔 question / answer1 / answer2"""
    items = []
    encoding = detect_encoding(filepath)
    with open(filepath, 'r', encoding=encoding) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = None
            for sep in ['\t', '|||']:
                if sep in line:
                    parts = line.split(sep)
                    break
            if parts and len(parts) >= 3:
                items.append({
                    'question': parts[0].strip(),
                    'answer1': parts[1].strip(),
                    'answer2': parts[2].strip(),
                })
    return items


def parse_file_similarity(filepath):
    """根据文件扩展名选择解析器（相似度对比模式）"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.csv':
        return parse_csv_similarity(filepath)
    elif ext == '.txt':
        return parse_txt_similarity(filepath)
    elif ext in ('.xlsx', '.xls'):
        return parse_excel_similarity(filepath)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")


# ==================== 双文件 QA 解析（每文件两列：Q + A）====================

import re as _re


def _find_qa_columns(header):
    """从表头中识别 question 和 answer 列索引"""
    q_idx, a_idx = None, None
    for i, name in enumerate(header):
        if q_idx is None and name in _Q_KEYWORDS:
            q_idx = i
        if a_idx is None and name in _A_GENERIC_KEYWORDS:
            a_idx = i
    return q_idx, a_idx


def parse_csv_qa(filepath):
    """解析 CSV 文件，提取 question / answer 两列"""
    items = []
    encoding = detect_encoding(filepath)
    with open(filepath, 'r', encoding=encoding) as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(f, dialect=dialect)
        rows = list(reader)

    if not rows:
        return items

    header = [name.strip().lower() for name in rows[0]]
    q_idx, a_idx = _find_qa_columns(header)

    if q_idx is not None and a_idx is not None:
        start = 1
    elif len(rows[0]) >= 2:
        q_idx, a_idx = 0, 1
        has_header = not _looks_like_data(rows[0][0]) if rows[0] else False
        start = 1 if has_header else 0
    else:
        return items

    for row in rows[start:]:
        if len(row) <= max(q_idx, a_idx):
            continue
        q = row[q_idx].strip() if row[q_idx] else ''
        a = row[a_idx].strip() if row[a_idx] else ''
        if not q:
            continue
        items.append({'question': q, 'answer': a})

    return items


def parse_excel_qa(filepath):
    """解析 Excel 文件，提取 question / answer 两列"""
    items = []
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return items

    header = [str(cell).strip().lower() if cell is not None else '' for cell in rows[0]]
    q_idx, a_idx = _find_qa_columns(header)

    if q_idx is not None and a_idx is not None:
        start = 1
    elif len(rows[0]) >= 2:
        q_idx, a_idx = 0, 1
        has_header = not _looks_like_data(str(rows[0][0])) if rows[0][0] is not None else False
        start = 1 if has_header else 0
    else:
        return items

    for row in rows[start:]:
        if len(row) <= max(q_idx, a_idx):
            continue
        q = str(row[q_idx]).strip() if row[q_idx] is not None else ''
        a = str(row[a_idx]).strip() if row[a_idx] is not None else ''
        if not q:
            continue
        items.append({'question': q, 'answer': a})

    return items


def parse_txt_qa(filepath):
    """解析 TXT 文件，每行用 TAB 或 ||| 分隔 question / answer"""
    items = []
    encoding = detect_encoding(filepath)
    with open(filepath, 'r', encoding=encoding) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = None
            for sep in ['\t', '|||']:
                if sep in line:
                    parts = line.split(sep, 1)
                    break
            if parts and len(parts) == 2:
                q = parts[0].strip()
                a = parts[1].strip()
                if q:
                    items.append({'question': q, 'answer': a})
    return items


def parse_file_qa(filepath):
    """根据文件扩展名选择解析器（QA 两列模式）"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.csv':
        return parse_csv_qa(filepath)
    elif ext == '.txt':
        return parse_txt_qa(filepath)
    elif ext in ('.xlsx', '.xls'):
        return parse_excel_qa(filepath)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")


def _normalize_question(q):
    """标准化问题文本用于匹配：strip + 压缩空白"""
    return _re.sub(r'\s+', ' ', q.strip())


def match_qa_files(items_a, items_b):
    """
    按问题精确匹配两个 QA 列表，返回 (matched_items, unmatched_details)。
    同一文件中重复问题取首条。
    """
    # 构建 file1 的 {normalized_q: (original_q, answer)} 字典
    dict_a = {}
    for item in items_a:
        key = _normalize_question(item['question'])
        if key and key not in dict_a:
            dict_a[key] = (item['question'], item['answer'])

    # 构建 file2 的字典
    dict_b = {}
    for item in items_b:
        key = _normalize_question(item['question'])
        if key and key not in dict_b:
            dict_b[key] = (item['question'], item['answer'])

    # 匹配
    matched = []
    keys_a = set(dict_a.keys())
    keys_b = set(dict_b.keys())
    common_keys = keys_a & keys_b

    for key in common_keys:
        q_a, a_a = dict_a[key]
        q_b, a_b = dict_b[key]
        matched.append({
            'question': q_a,
            'answer1': a_a,
            'answer2': a_b,
        })

    # 未匹配
    only_a = [dict_a[k][0] for k in (keys_a - keys_b)]
    only_b = [dict_b[k][0] for k in (keys_b - keys_a)]

    unmatched = {
        'only_in_file1': only_a,
        'only_in_file2': only_b,
        'total': len(only_a) + len(only_b),
    }

    return matched, unmatched

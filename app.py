"""语料评测平台 - 主应用"""
import os
import json
import uuid
import logging
import traceback
import requests as req_lib
from flask import Flask, render_template, request, jsonify, redirect, url_for, flash

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)
from werkzeug.utils import secure_filename

from config import UPLOAD_FOLDER, ALLOWED_EXTENSIONS, MAX_CONTENT_LENGTH, DATABASE_URI, DOUBAO_PROFILE_DIR, DOUBAO_LOGIN_TIMEOUT, LLM_CONCURRENCY
from models import db, LLMConfig, CorpusFile, CorpusItem, PromptTemplate, SimilarityTask, SimilarityItem, DoubaoTask, DoubaoItem
from services.file_parser import parse_file, parse_file_similarity, parse_file_qa, match_qa_files, allowed_file
from services.llm_service import LLMService
from services.task_executor import execute_parallel_batch
from services.classifier import (
    classify_subjective_objective, classify_difficulty,
    classify_category, generate_objective_answer,
    evaluate_quality, classify_domain, classify_intent,
    classify_combined,
    DEFAULT_CLASSIFY_SUBJ_OBJ_PROMPT, DEFAULT_CLASSIFY_DIFFICULTY_PROMPT,
    DEFAULT_CLASSIFY_CATEGORY_PROMPT, DEFAULT_GENERATE_ANSWER_PROMPT,
    DEFAULT_QUALITY_EVAL_PROMPT, DEFAULT_DOMAIN_CLASSIFY_PROMPT,
    DEFAULT_INTENT_CLASSIFY_PROMPT, DEFAULT_COMBINED_CLASSIFY_PROMPT,
)
from services.scorer import score_answer, clean_answer_v2, DEFAULT_SCORE_PROMPT
from services.comparator import compare_similarity, DEFAULT_SIMILARITY_COMPARE_PROMPT
from services.doubao_auth import DoubaoAuthManager

# 豆包认证管理器（单例）
doubao_auth = DoubaoAuthManager(user_data_dir=DOUBAO_PROFILE_DIR, login_timeout=DOUBAO_LOGIN_TIMEOUT)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'corpus-eval-platform-secret'
app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URI
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

db.init_app(app)


# API 请求返回 JSON 格式的错误，避免前端 fetch 解析 HTML 报错
@app.errorhandler(413)
def request_entity_too_large(e):
    return jsonify({'error': '文件太大，最大支持 50MB'}), 413


@app.errorhandler(400)
def bad_request(e):
    return jsonify({'error': f'请求错误: {e.description}'}), 400


@app.errorhandler(500)
def internal_error(e):
    return jsonify({'error': '服务器内部错误，请稍后重试'}), 500

def _auto_migrate(engine):
    """自动检测并添加缺失的数据库列，避免因 schema 变更需要删库重建"""
    from sqlalchemy import inspect as sa_inspect, text
    inspector = sa_inspect(engine)
    with engine.connect() as conn:
        for table_name, table_obj in db.metadata.tables.items():
            if not inspector.has_table(table_name):
                continue
            existing_cols = {col['name'] for col in inspector.get_columns(table_name)}
            for col in table_obj.columns:
                if col.name not in existing_cols:
                    col_type = col.type.compile(engine.dialect)
                    conn.execute(text(f'ALTER TABLE {table_name} ADD COLUMN {col.name} {col_type}'))
        conn.commit()


with app.app_context():
    os.makedirs(os.path.dirname(DATABASE_URI.replace('sqlite:///', '')), exist_ok=True)
    db.create_all()
    _auto_migrate(db.engine)
    # 初始化 / 同步默认 prompt 模板（代码更新后 DB 自动同步）
    defaults = [
        ('主观/客观分类', 'classify_subj_obj', DEFAULT_CLASSIFY_SUBJ_OBJ_PROMPT),
        ('难度等级分类', 'classify_difficulty', DEFAULT_CLASSIFY_DIFFICULTY_PROMPT),
        ('自定义分类', 'classify_category', DEFAULT_CLASSIFY_CATEGORY_PROMPT),
        ('客观题答案生成', 'generate_answer', DEFAULT_GENERATE_ANSWER_PROMPT),
        ('Answer 打分', 'score_answer', DEFAULT_SCORE_PROMPT),
        ('语料质量评估', 'quality_eval', DEFAULT_QUALITY_EVAL_PROMPT),
        ('领域/场景分类', 'domain_classify', DEFAULT_DOMAIN_CLASSIFY_PROMPT),
        ('意图识别分类', 'intent_classify', DEFAULT_INTENT_CLASSIFY_PROMPT),
        ('综合分类(一次调用)', 'combined_classify', DEFAULT_COMBINED_CLASSIFY_PROMPT),
        ('答案相似度对比', 'similarity_compare', DEFAULT_SIMILARITY_COMPARE_PROMPT),
    ]
    for name, ptype, content in defaults:
        existing = PromptTemplate.query.filter_by(prompt_type=ptype, is_default=True).first()
        if not existing:
            tpl = PromptTemplate(name=name, prompt_type=ptype, content=content, is_default=True)
            db.session.add(tpl)
        else:
            # 同步更新默认模板内容（仅当用户未自行修改过时）
            existing.content = content
            existing.name = name
    db.session.commit()


def get_llm_service(config_id=None):
    """获取 LLM 服务实例"""
    if config_id:
        cfg = LLMConfig.query.get(config_id)
    else:
        cfg = LLMConfig.query.filter_by(is_default=True).first()
    if not cfg:
        cfg = LLMConfig.query.first()
    if not cfg:
        return None
    return LLMService(cfg.api_url, cfg.api_key, cfg.model, cfg.proxy, cfg.verify_ssl)


# ==================== 页面路由 ====================

@app.route('/favicon.ico')
def favicon():
    return '', 204  # No Content


@app.route('/')
def index():
    files = CorpusFile.query.order_by(CorpusFile.created_at.desc()).all()
    configs = LLMConfig.query.all()
    return render_template('index.html', files=files, configs=configs)


@app.route('/config')
def config_page():
    configs = LLMConfig.query.order_by(LLMConfig.created_at.desc()).all()
    prompts = PromptTemplate.query.order_by(PromptTemplate.prompt_type).all()
    return render_template('config.html', configs=configs, prompts=prompts)


@app.route('/corpus/<int:file_id>')
def corpus_detail(file_id):
    corpus = CorpusFile.query.get_or_404(file_id)
    items = CorpusItem.query.filter_by(file_id=file_id).all()
    configs = LLMConfig.query.all()
    prompts = PromptTemplate.query.all()
    return render_template('corpus_detail.html', corpus=corpus, items=items,
                           configs=configs, prompts=prompts)


@app.route('/similarity')
def similarity_page():
    tasks = SimilarityTask.query.order_by(SimilarityTask.created_at.desc()).all()
    configs = LLMConfig.query.all()
    return render_template('similarity.html', tasks=tasks, configs=configs)


@app.route('/similarity/<int:task_id>')
def similarity_detail(task_id):
    task = SimilarityTask.query.get_or_404(task_id)
    items = SimilarityItem.query.filter_by(task_id=task_id).all()
    configs = LLMConfig.query.all()
    prompts = PromptTemplate.query.all()
    return render_template('similarity_detail.html', task=task, items=items,
                           configs=configs, prompts=prompts)


# ==================== API 路由 ====================

# ----- LLM 配置 -----
@app.route('/api/llm-config', methods=['POST'])
def save_llm_config():
    data = request.json
    name = data.get('name', '').strip()
    api_url = data.get('api_url', '').strip()
    api_key = data.get('api_key', '').strip()
    model = data.get('model', '').strip()
    proxy = (data.get('proxy') or '').strip() or None
    verify_ssl = data.get('verify_ssl', True)
    is_default = data.get('is_default', False)

    if not all([name, api_url, api_key, model]):
        return jsonify({'error': '所有字段都是必填的'}), 400

    if is_default:
        LLMConfig.query.update({'is_default': False})

    config = LLMConfig(
        name=name, 
        api_url=api_url, 
        api_key=api_key,
        model=model, 
        proxy=proxy,
        verify_ssl=verify_ssl,
        is_default=is_default
    )
    db.session.add(config)
    db.session.commit()
    return jsonify({'id': config.id, 'message': '保存成功'})


@app.route('/api/llm-config/<int:config_id>', methods=['DELETE'])
def delete_llm_config(config_id):
    config = LLMConfig.query.get_or_404(config_id)
    db.session.delete(config)
    db.session.commit()
    return jsonify({'message': '删除成功'})


@app.route('/api/llm-config/<int:config_id>/test', methods=['POST'])
def test_llm_config(config_id):
    config = LLMConfig.query.get_or_404(config_id)
    try:
        llm = LLMService(config.api_url, config.api_key, config.model)
        result = llm.chat("你是一个助手。", "请回复'连接成功'这四个字。", temperature=0)
        return jsonify({'message': '连接测试成功', 'response': result})
    except req_lib.exceptions.Timeout:
        return jsonify({'error': '连接超时，请检查网络或增大超时时间'}), 500
    except req_lib.exceptions.ConnectionError as e:
        return jsonify({'error': f'网络连接失败: {str(e)}'}), 500
    except Exception as e:
        # 记录详细错误日志
        import traceback
        print(f"[LLM Test Error] Config ID {config_id}: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': f'连接失败: {str(e)}'}), 500


# ----- Prompt 模板 -----
@app.route('/api/prompt', methods=['POST'])
def save_prompt():
    data = request.json
    name = data.get('name', '').strip()
    prompt_type = data.get('prompt_type', '').strip()
    content = data.get('content', '').strip()

    if not all([name, prompt_type, content]):
        return jsonify({'error': '所有字段都是必填的'}), 400

    prompt = PromptTemplate(name=name, prompt_type=prompt_type, content=content)
    db.session.add(prompt)
    db.session.commit()
    return jsonify({'id': prompt.id, 'message': '保存成功'})


@app.route('/api/prompt/<int:prompt_id>', methods=['PUT'])
def update_prompt(prompt_id):
    prompt = PromptTemplate.query.get_or_404(prompt_id)
    data = request.json
    prompt.content = data.get('content', prompt.content)
    prompt.name = data.get('name', prompt.name)
    db.session.commit()
    return jsonify({'message': '更新成功'})


@app.route('/api/prompt/<int:prompt_id>', methods=['DELETE'])
def delete_prompt(prompt_id):
    prompt = PromptTemplate.query.get_or_404(prompt_id)
    if prompt.is_default:
        return jsonify({'error': '不能删除默认模板'}), 400
    db.session.delete(prompt)
    db.session.commit()
    return jsonify({'message': '删除成功'})


# ----- 文件上传 -----
@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': '没有选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400

    corpus_type = request.form.get('corpus_type', 'question')

    if not allowed_file(file.filename, ALLOWED_EXTENSIONS):
        return jsonify({'error': f'不支持的文件格式，仅支持: {", ".join(ALLOWED_EXTENSIONS)}'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if not ext:
        return jsonify({'error': '无法识别文件扩展名'}), 400
    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        items = parse_file(filepath, corpus_type)
        if not items:
            os.remove(filepath)
            return jsonify({'error': '文件解析失败或文件为空'}), 400

        corpus = CorpusFile(
            filename=filename,
            original_name=file.filename,
            file_type=ext,
            corpus_type=corpus_type,
            record_count=len(items),
        )
        db.session.add(corpus)
        db.session.flush()

        for item_data in items:
            item = CorpusItem(
                file_id=corpus.id,
                question=item_data['question'],
                answer=item_data.get('answer', ''),
            )
            db.session.add(item)

        db.session.commit()
        return jsonify({
            'id': corpus.id,
            'message': f'上传成功，解析到 {len(items)} 条语料',
            'count': len(items),
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'解析失败: {str(e)}'}), 500


@app.route('/api/corpus/<int:file_id>', methods=['DELETE'])
def delete_corpus(file_id):
    corpus = CorpusFile.query.get_or_404(file_id)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], corpus.filename)
    if os.path.exists(filepath):
        os.remove(filepath)
    db.session.delete(corpus)
    db.session.commit()
    return jsonify({'message': '删除成功'})


# ----- 分类与评测 -----
@app.route('/api/classify-stream', methods=['POST'])
def classify_items_stream():
    """流式分类语料（支持并行调用、重试、暂停恢复）"""
    from flask import Response, stream_with_context

    # 在生成器外读取请求数据，避免请求上下文问题
    data = request.get_json(force=True, silent=True) or {}
    file_id = data.get('file_id')
    config_id = data.get('config_id')
    classify_type = data.get('classify_type')
    classify_types = data.get('classify_types', [])
    categories = data.get('categories', '')
    prompt_id = data.get('prompt_id')
    item_ids = data.get('item_ids', [])
    try:
        concurrency = int(data.get('concurrency', LLM_CONCURRENCY))
    except (TypeError, ValueError):
        concurrency = LLM_CONCURRENCY

    logger.info(f"[分类流] 开始: classify_type={classify_type}, classify_types={classify_types}, "
                f"config_id={config_id}, prompt_id={prompt_id}, file_id={file_id}, "
                f"item_ids数量={len(item_ids)}, 并行数={concurrency}")

    def generate():
      try:
        llm = get_llm_service(config_id)
        if not llm:
            logger.error("[分类流] 未找到模型配置")
            yield f"data: {json.dumps({'error': '请先配置大模型'})}\n\n"
            return

        logger.info(f"[分类流] 使用模型: {llm.model}")

        custom_prompt = None
        if prompt_id:
            tpl = PromptTemplate.query.get(prompt_id)
            if tpl:
                custom_prompt = tpl.content
                logger.info(f"[分类流] 使用自定义 Prompt: {tpl.name} (type={tpl.prompt_type})")

        if item_ids:
            items = CorpusItem.query.filter(CorpusItem.id.in_(item_ids)).all()
        else:
            items = CorpusItem.query.filter_by(file_id=file_id).all()

        total = len(items)
        logger.info(f"[分类流] 待处理条目: {total} 条")

        if total == 0:
            yield f"data: {json.dumps({'error': '没有找到待处理的条目'})}\n\n"
            return

        # 发送 init 事件，供前端暂停恢复使用
        item_id_list = [item.id for item in items]
        yield f"data: {json.dumps({'status': 'init', 'total': total, 'item_ids': item_id_list})}\n\n"

        # 预缓存 ORM 属性到普通 Python 属性，避免 commit 后线程不安全的延迟加载
        for item in items:
            item._q = item.question

        def do_classify(item):
            """纯 LLM 调用（在子线程中执行，不操作数据库）"""
            q = item._q
            if classify_type == 'combined':
                return ('combined', classify_combined(llm, q, categories, custom_prompt))
            elif classify_type == 'subj_obj':
                res = classify_subjective_objective(llm, q, custom_prompt)
                if res.get('type') == 'objective':
                    try:
                        ans_res = generate_objective_answer(llm, q)
                        res['_objective_answer'] = ans_res.get('answer', '')
                    except Exception:
                        pass
                return ('subj_obj', res)
            elif classify_type == 'difficulty':
                return ('difficulty', classify_difficulty(llm, q, custom_prompt))
            elif classify_type == 'category':
                return ('category', classify_category(llm, q, categories, custom_prompt))
            elif classify_type == 'quality':
                return ('quality', evaluate_quality(llm, q, custom_prompt))
            elif classify_type == 'domain':
                return ('domain', classify_domain(llm, q, custom_prompt))
            elif classify_type == 'intent':
                return ('intent', classify_intent(llm, q, custom_prompt))
            else:
                raise ValueError(f'未知分类类型: {classify_type}')

        for idx, item, result, error in execute_parallel_batch(items, do_classify, concurrency):
            if error:
                db.session.rollback()
                logger.error(f"[分类流] 第 {idx}/{total} 条处理失败 (item_id={item.id}): {str(error)}")
                logger.error(''.join(traceback.format_exception(type(error), error, error.__traceback__)))
                yield f"data: {json.dumps({'status': 'error', 'item_id': item.id, 'error': str(error), 'progress': idx, 'total': total})}\n\n"
            else:
                ctype, res = result
                # 在主线程中写入数据库
                if ctype == 'combined':
                    _apply_combined_result(item, res, classify_types, categories)
                elif ctype == 'subj_obj':
                    item.subj_obj = res.get('type', '')
                    if res.get('_objective_answer'):
                        item.objective_answer = res['_objective_answer']
                elif ctype == 'difficulty':
                    item.difficulty = res.get('difficulty', '')
                elif ctype == 'category':
                    item.category = res.get('category', '')
                elif ctype == 'quality':
                    item.quality_score = float(res.get('quality_score', 0))
                    item.quality_label = res.get('quality_label', '')
                    item.quality_detail = json.dumps(res, ensure_ascii=False)
                elif ctype == 'domain':
                    item.domain = res.get('domain', '')
                    item.sub_domain = res.get('sub_domain', '')
                elif ctype == 'intent':
                    item.intent = res.get('intent', '')
                    item.intent_cn = res.get('intent_cn', '')
                    item.intent_confidence = float(res.get('confidence', 0))

                db.session.commit()
                logger.info(f"[分类流] 第 {idx}/{total} 条处理成功")
                yield f"data: {json.dumps({'status': 'ok', 'item_id': item.id, 'progress': idx, 'total': total, 'data': res})}\n\n"

        yield f"data: {json.dumps({'status': 'done', 'message': f'分类完成，共 {total} 条'})}\n\n"
      except Exception as outer_err:
        logger.error(f"[分类流] 生成器异常: {str(outer_err)}")
        logger.error(traceback.format_exc())
        yield f"data: {json.dumps({'error': f'分类异常: {str(outer_err)}'})}\n\n"

    return Response(stream_with_context(generate()), mimetype='text/event-stream')


@app.route('/api/classify', methods=['POST'])
def classify_items():
    """批量分类语料"""
    data = request.json
    file_id = data.get('file_id')
    config_id = data.get('config_id')
    classify_type = data.get('classify_type')  # subj_obj / difficulty / category / quality / domain / intent / combined
    classify_types = data.get('classify_types', [])  # combined 模式下指定哪些维度
    categories = data.get('categories', '')  # 自定义分类类型（逗号分隔）
    prompt_id = data.get('prompt_id')
    item_ids = data.get('item_ids', [])  # 可选，指定条目

    llm = get_llm_service(config_id)
    if not llm:
        return jsonify({'error': '请先配置大模型'}), 400

    custom_prompt = None
    if prompt_id:
        tpl = PromptTemplate.query.get(prompt_id)
        if tpl:
            custom_prompt = tpl.content

    if item_ids:
        items = CorpusItem.query.filter(CorpusItem.id.in_(item_ids)).all()
    else:
        items = CorpusItem.query.filter_by(file_id=file_id).all()

    results = []
    errors = []
    for item in items:
        try:
            if classify_type == 'combined':
                # 一次调用返回所有维度
                res = classify_combined(llm, item.question, categories, custom_prompt)
                _apply_combined_result(item, res, classify_types, categories)
                results.append({'id': item.id, 'status': 'ok'})

            elif classify_type == 'subj_obj':
                res = classify_subjective_objective(llm, item.question, custom_prompt)
                item.subj_obj = res.get('type', '')
                if item.subj_obj == 'objective':
                    try:
                        ans_res = generate_objective_answer(llm, item.question)
                        item.objective_answer = ans_res.get('answer', '')
                    except Exception:
                        pass
                results.append({'id': item.id, 'subj_obj': item.subj_obj})

            elif classify_type == 'difficulty':
                res = classify_difficulty(llm, item.question, custom_prompt)
                item.difficulty = res.get('difficulty', '')
                results.append({'id': item.id, 'difficulty': item.difficulty})

            elif classify_type == 'category':
                if not categories:
                    return jsonify({'error': '自定义分类需要提供分类类型'}), 400
                res = classify_category(llm, item.question, categories, custom_prompt)
                item.category = res.get('category', '')
                results.append({'id': item.id, 'category': item.category})

            elif classify_type == 'quality':
                res = evaluate_quality(llm, item.question, custom_prompt)
                item.quality_score = float(res.get('quality_score', 0))
                item.quality_label = res.get('quality_label', '')
                item.quality_detail = json.dumps(res, ensure_ascii=False)
                results.append({'id': item.id, 'quality_score': item.quality_score})

            elif classify_type == 'domain':
                res = classify_domain(llm, item.question, custom_prompt)
                item.domain = res.get('domain', '')
                item.sub_domain = res.get('sub_domain', '')
                results.append({'id': item.id, 'domain': item.domain})

            elif classify_type == 'intent':
                res = classify_intent(llm, item.question, custom_prompt)
                item.intent = res.get('intent', '')
                item.intent_cn = res.get('intent_cn', '')
                item.intent_confidence = float(res.get('confidence', 0))
                results.append({'id': item.id, 'intent': item.intent})

        except Exception as e:
            errors.append({'id': item.id, 'question': item.question[:50], 'error': str(e)})

    db.session.commit()
    return jsonify({
        'message': f'分类完成: 成功 {len(results)} 条, 失败 {len(errors)} 条',
        'results': results,
        'errors': errors,
    })


def _apply_combined_result(item, res, classify_types, categories):
    """将综合分类结果写入 CorpusItem 对应字段"""
    types = set(classify_types) if classify_types else {'subj_obj', 'difficulty', 'quality', 'domain', 'intent', 'category'}

    if 'subj_obj' in types:
        item.subj_obj = res.get('subj_obj', '')
        obj_answer = res.get('objective_answer', '')
        if item.subj_obj == 'objective' and obj_answer:
            item.objective_answer = obj_answer

    if 'difficulty' in types:
        item.difficulty = res.get('difficulty', '')

    if 'quality' in types:
        item.quality_score = float(res.get('quality_score', 0))
        item.quality_label = res.get('quality_label', '')
        item.quality_detail = json.dumps(res, ensure_ascii=False)

    if 'domain' in types:
        item.domain = res.get('domain', '')
        item.sub_domain = res.get('sub_domain', '')

    if 'intent' in types:
        item.intent = res.get('intent', '')
        item.intent_cn = res.get('intent_cn', '')
        item.intent_confidence = float(res.get('intent_confidence', res.get('confidence', 0)))

    if 'category' in types and categories:
        item.category = res.get('category', '')


@app.route('/api/score', methods=['POST'])
def score_items():
    """批量对 Answer 打分"""
    data = request.json
    file_id = data.get('file_id')
    config_id = data.get('config_id')
    prompt_id = data.get('prompt_id')
    item_ids = data.get('item_ids', [])

    llm = get_llm_service(config_id)
    if not llm:
        return jsonify({'error': '请先配置大模型'}), 400

    custom_prompt = None
    if prompt_id:
        tpl = PromptTemplate.query.get(prompt_id)
        if tpl:
            custom_prompt = tpl.content

    if item_ids:
        items = CorpusItem.query.filter(CorpusItem.id.in_(item_ids)).all()
    else:
        items = CorpusItem.query.filter_by(file_id=file_id).all()

    results = []
    errors = []
    for item in items:
        if not item.answer:
            continue
        try:
            res = score_answer(llm, item.question, item.answer, custom_prompt)
            item.answer_score = float(res.get('score', 0))
            item.score_reason = res.get('reason', '')
            results.append({
                'id': item.id,
                'score': item.answer_score,
                'reason': item.score_reason,
            })
        except Exception as e:
            errors.append({'id': item.id, 'question': item.question[:50], 'error': str(e)})

    db.session.commit()
    return jsonify({
        'message': f'打分完成: 成功 {len(results)} 条, 失败 {len(errors)} 条',
        'results': results,
        'errors': errors,
    })


@app.route('/api/score-stream', methods=['POST'])
def score_items_stream():
    """流式打分（支持并行调用、重试、暂停恢复）"""
    from flask import Response, stream_with_context

    data = request.get_json(force=True, silent=True) or {}
    file_id = data.get('file_id')
    config_id = data.get('config_id')
    prompt_id = data.get('prompt_id')
    item_ids = data.get('item_ids', [])
    try:
        concurrency = int(data.get('concurrency', LLM_CONCURRENCY))
    except (TypeError, ValueError):
        concurrency = LLM_CONCURRENCY

    logger.info(f"[打分流] 开始: config_id={config_id}, prompt_id={prompt_id}, "
                f"file_id={file_id}, item_ids数量={len(item_ids)}, 并行数={concurrency}")

    def generate():
      try:
        llm = get_llm_service(config_id)
        if not llm:
            yield f"data: {json.dumps({'error': '请先配置大模型'})}\n\n"
            return

        custom_prompt = None
        if prompt_id:
            tpl = PromptTemplate.query.get(prompt_id)
            if tpl:
                custom_prompt = tpl.content

        if item_ids:
            items = CorpusItem.query.filter(CorpusItem.id.in_(item_ids)).all()
        else:
            items = CorpusItem.query.filter_by(file_id=file_id).all()

        # 过滤掉没有 answer 的条目
        scorable_items = [item for item in items if item.answer]
        total = len(scorable_items)
        logger.info(f"[打分流] 待打分条目: {total} 条")

        if total == 0:
            yield f"data: {json.dumps({'error': '没有找到有答案的条目'})}\n\n"
            return

        # 发送 init 事件
        item_id_list = [item.id for item in scorable_items]
        yield f"data: {json.dumps({'status': 'init', 'total': total, 'item_ids': item_id_list})}\n\n"

        # 预缓存 ORM 属性到普通 Python 属性，避免 commit 后线程不安全的延迟加载
        for item in scorable_items:
            item._q = item.question
            item._a = item.answer

        def do_score(item):
            """纯 LLM 调用（在子线程中执行）"""
            return score_answer(llm, item._q, item._a, custom_prompt)

        for idx, item, result, error in execute_parallel_batch(scorable_items, do_score, concurrency):
            if error:
                db.session.rollback()
                logger.error(f"[打分流] 第 {idx}/{total} 条处理失败 (item_id={item.id}): {str(error)}")
                logger.error(''.join(traceback.format_exception(type(error), error, error.__traceback__)))
                yield f"data: {json.dumps({'status': 'error', 'item_id': item.id, 'error': str(error), 'progress': idx, 'total': total})}\n\n"
            else:
                item.answer_score = float(result.get('score', 0))
                item.score_reason = result.get('reason', '')
                db.session.commit()
                logger.info(f"[打分流] 第 {idx}/{total} 条处理成功")
                yield f"data: {json.dumps({'status': 'ok', 'item_id': item.id, 'progress': idx, 'total': total, 'data': {'score': item.answer_score, 'reason': item.score_reason}})}\n\n"

        yield f"data: {json.dumps({'status': 'done', 'message': f'打分完成，共 {total} 条'})}\n\n"
      except Exception as outer_err:
        logger.error(f"[打分流] 生成器异常: {str(outer_err)}")
        logger.error(traceback.format_exc())
        yield f"data: {json.dumps({'error': f'打分异常: {str(outer_err)}'})}\n\n"

    return Response(stream_with_context(generate()), mimetype='text/event-stream')


@app.route('/api/export/<int:file_id>')
def export_results(file_id):
    """导出评测结果为 Excel"""
    import io
    from flask import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    corpus = CorpusFile.query.get_or_404(file_id)
    items = CorpusItem.query.filter_by(file_id=file_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '评测结果'

    # 表头
    headers = ['ID', '问题', '答案', '主观/客观', '难度', '分类',
               '客观题答案', '质量评分', '质量等级', '领域', '子领域',
               '意图', '意图(中文)', '意图置信度', '答案评分', '评分理由']
    ws.append(headers)

    # 表头样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='344955', end_color='344955', fill_type='solid')
    thin_border = Border(
        bottom=Side(style='thin', color='CCCCCC'),
    )
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 数据行
    for idx, item in enumerate(items, 1):
        difficulty_display = ''
        if item.difficulty == 'L1':
            difficulty_display = 'L1(简单)'
        elif item.difficulty == 'L2':
            difficulty_display = 'L2(中等)'
        elif item.difficulty == 'L3':
            difficulty_display = 'L3(困难)'
        elif item.difficulty:
            difficulty_display = item.difficulty

        subj_obj_display = ''
        if item.subj_obj == 'objective':
            subj_obj_display = '客观'
        elif item.subj_obj == 'subjective':
            subj_obj_display = '主观'

        ws.append([
            idx, item.question, item.answer or '',
            subj_obj_display, difficulty_display, item.category or '',
            item.objective_answer or '',
            item.quality_score or '', item.quality_label or '',
            item.domain or '', item.sub_domain or '',
            item.intent or '', item.intent_cn or '',
            item.intent_confidence or '',
            item.answer_score or '', item.score_reason or '',
        ])

    # 数据行边框
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    # 列宽
    col_widths = [6, 40, 30, 10, 12, 12, 30, 10, 10, 12, 12, 18, 12, 10, 10, 30]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    filename = os.path.splitext(corpus.original_name)[0] + '_评测结果.xlsx'
    encoded = quote(filename)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded}"}
    )


# ==================== 相似度对比 API ====================

@app.route('/api/upload-similarity', methods=['POST'])
def upload_similarity():
    """上传相似度对比文件"""
    if 'file' not in request.files:
        return jsonify({'error': '没有选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400

    if not allowed_file(file.filename, ALLOWED_EXTENSIONS):
        return jsonify({'error': f'不支持的文件格式，仅支持: {", ".join(ALLOWED_EXTENSIONS)}'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if not ext:
        return jsonify({'error': '无法识别文件扩展名'}), 400
    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        items = parse_file_similarity(filepath)
        if not items:
            os.remove(filepath)
            return jsonify({'error': '文件解析失败或文件为空，请确保包含 question/answer1/answer2 三列'}), 400

        task = SimilarityTask(
            filename=filename,
            original_name=file.filename,
            file_type=ext,
            record_count=len(items),
        )
        db.session.add(task)
        db.session.flush()

        for item_data in items:
            a1 = item_data.get('answer1', '')
            a2 = item_data.get('answer2', '')
            item = SimilarityItem(
                task_id=task.id,
                question=item_data['question'],
                answer1=a1,
                answer2=a2,
                answer1_cleaned=clean_answer_v2(a1),
                answer2_cleaned=clean_answer_v2(a2),
            )
            db.session.add(item)

        db.session.commit()
        return jsonify({
            'id': task.id,
            'message': f'上传成功，解析到 {len(items)} 条对比数据',
            'count': len(items),
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'解析失败: {str(e)}'}), 500


@app.route('/api/upload-similarity-dual', methods=['POST'])
def upload_similarity_dual():
    """上传两个 QA 文件进行相似度对比"""
    if 'file1' not in request.files or 'file2' not in request.files:
        return jsonify({'error': '请同时选择两个文件'}), 400

    file1 = request.files['file1']
    file2 = request.files['file2']
    if file1.filename == '' or file2.filename == '':
        return jsonify({'error': '请同时选择两个文件'}), 400

    for f in [file1, file2]:
        if not allowed_file(f.filename, ALLOWED_EXTENSIONS):
            return jsonify({'error': f'不支持的文件格式（{f.filename}），仅支持: {", ".join(ALLOWED_EXTENSIONS)}'}), 400

    # 保存两个文件
    ext1 = file1.filename.rsplit('.', 1)[-1].lower() if '.' in file1.filename else ''
    ext2 = file2.filename.rsplit('.', 1)[-1].lower() if '.' in file2.filename else ''
    fname1 = f"{uuid.uuid4().hex}.{ext1}"
    fname2 = f"{uuid.uuid4().hex}.{ext2}"
    fpath1 = os.path.join(app.config['UPLOAD_FOLDER'], fname1)
    fpath2 = os.path.join(app.config['UPLOAD_FOLDER'], fname2)
    file1.save(fpath1)
    file2.save(fpath2)

    try:
        items_a = parse_file_qa(fpath1)
        if not items_a:
            raise ValueError(f'文件1（{file1.filename}）解析失败或为空，请确保包含 Q 和 A 两列')

        items_b = parse_file_qa(fpath2)
        if not items_b:
            raise ValueError(f'文件2（{file2.filename}）解析失败或为空，请确保包含 Q 和 A 两列')

        matched, unmatched = match_qa_files(items_a, items_b)
        if not matched:
            raise ValueError(f'两个文件中没有匹配的问题（文件1: {len(items_a)} 条，文件2: {len(items_b)} 条）')

        task = SimilarityTask(
            filename=fname1,
            original_name=f"{file1.filename} ↔ {file2.filename}",
            file_type=ext1,
            record_count=len(matched),
            upload_mode='dual',
            filename2=fname2,
            original_name2=file2.filename,
            unmatched_count=unmatched['total'],
            unmatched_json=json.dumps(unmatched, ensure_ascii=False) if unmatched['total'] > 0 else None,
        )
        db.session.add(task)
        db.session.flush()

        for item_data in matched:
            a1 = item_data['answer1']
            a2 = item_data['answer2']
            item = SimilarityItem(
                task_id=task.id,
                question=item_data['question'],
                answer1=a1,
                answer2=a2,
                answer1_cleaned=clean_answer_v2(a1),
                answer2_cleaned=clean_answer_v2(a2),
            )
            db.session.add(item)

        db.session.commit()

        resp = {
            'id': task.id,
            'message': f'上传成功，匹配到 {len(matched)} 条对比数据',
            'count': len(matched),
        }
        if unmatched['total'] > 0:
            resp['unmatched'] = {
                'total': unmatched['total'],
                'only_in_file1': len(unmatched['only_in_file1']),
                'only_in_file2': len(unmatched['only_in_file2']),
            }
            resp['message'] += f'（{unmatched["total"]} 个问题未匹配）'
        return jsonify(resp)

    except Exception as e:
        db.session.rollback()
        for fp in [fpath1, fpath2]:
            if os.path.exists(fp):
                try:
                    os.remove(fp)
                except OSError:
                    pass
        return jsonify({'error': str(e)}), 400


@app.route('/api/similarity/<int:task_id>', methods=['DELETE'])
def delete_similarity(task_id):
    """删除相似度对比任务"""
    task = SimilarityTask.query.get_or_404(task_id)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], task.filename)
    if os.path.exists(filepath):
        os.remove(filepath)
    if task.filename2:
        filepath2 = os.path.join(app.config['UPLOAD_FOLDER'], task.filename2)
        if os.path.exists(filepath2):
            os.remove(filepath2)
    db.session.delete(task)
    db.session.commit()
    return jsonify({'message': '删除成功'})


@app.route('/api/similarity-compare-stream', methods=['POST'])
def similarity_compare_stream():
    """流式相似度对比（支持并行调用、重试、暂停恢复）"""
    from flask import Response, stream_with_context

    # 在生成器外部预先解析请求数据，避免生成器内部读取不到请求体
    data = request.get_json(force=True, silent=True) or {}
    task_id = data.get('task_id')
    config_id = data.get('config_id')
    prompt_id = data.get('prompt_id')
    item_ids = data.get('item_ids', [])
    try:
        concurrency = int(data.get('concurrency', LLM_CONCURRENCY))
    except (TypeError, ValueError):
        concurrency = LLM_CONCURRENCY

    def generate():
      try:
        llm = get_llm_service(config_id)
        if not llm:
            yield f"data: {json.dumps({'error': '请先配置大模型'})}\n\n"
            return

        custom_prompt = None
        if prompt_id:
            tpl = PromptTemplate.query.get(prompt_id)
            if tpl:
                custom_prompt = tpl.content

        task = SimilarityTask.query.get(task_id)
        if not task:
            yield f"data: {json.dumps({'error': '任务不存在'})}\n\n"
            return

        task.status = 'running'
        db.session.commit()

        if item_ids:
            items = SimilarityItem.query.filter(SimilarityItem.id.in_(item_ids)).all()
        else:
            items = SimilarityItem.query.filter_by(task_id=task_id).all()

        # 预先过滤掉空答案的条目，记录 skip
        processable_items = []
        skip_count = 0
        for item in items:
            a1 = item.answer1_cleaned or clean_answer_v2(item.answer1 or '')
            a2 = item.answer2_cleaned or clean_answer_v2(item.answer2 or '')
            if not a1.strip() and not a2.strip():
                skip_count += 1
                continue
            # 缓存清洗后的答案供子线程使用
            item._clean_a1 = a1
            item._clean_a2 = a2
            # 预缓存 ORM 属性，避免 commit 后线程不安全的延迟加载
            item._q = item.question
            processable_items.append(item)

        total = len(items)
        processable_total = len(processable_items)

        # 发送 init 事件（仅包含实际需要处理的条目，不含 skip 的）
        item_id_list = [item.id for item in processable_items]
        yield f"data: {json.dumps({'status': 'init', 'total': processable_total, 'item_ids': item_id_list})}\n\n"

        # 发送 skip 事件
        if skip_count > 0:
            logger.info(f"[相似度对比] 跳过 {skip_count} 条空答案条目")

        def do_compare(item):
            """纯 LLM 调用（在子线程中执行）"""
            return compare_similarity(llm, item._q, item._clean_a1, item._clean_a2, custom_prompt)

        completed = 0

        for idx, item, result, error in execute_parallel_batch(processable_items, do_compare, concurrency):
            if error:
                db.session.rollback()
                logger.error(f"[相似度对比] 第 {idx}/{processable_total} 条处理失败 (item_id={item.id}): {str(error)}")
                logger.error(''.join(traceback.format_exception(type(error), error, error.__traceback__)))
                yield f"data: {json.dumps({'status': 'error', 'item_id': item.id, 'error': str(error), 'progress': idx, 'total': processable_total})}\n\n"
            else:
                item.similarity_score = float(result.get('similarity_score', 0))
                item.similarity_label = result.get('similarity_label', '')
                kd = result.get('key_differences', '')
                if isinstance(kd, list):
                    kd = '；'.join(str(x) for x in kd)
                item.key_differences = str(kd) if kd else ''
                item.detail_json = json.dumps(result, ensure_ascii=False)
                completed += 1
                task.compared_count = completed
                db.session.commit()

                yield f"data: {json.dumps({'status': 'ok', 'item_id': item.id, 'progress': idx, 'total': processable_total, 'data': {'similarity_score': item.similarity_score, 'similarity_label': item.similarity_label, 'key_differences': item.key_differences}})}\n\n"

        task.status = 'completed'
        db.session.commit()
        yield f"data: {json.dumps({'status': 'done', 'message': f'对比完成，共 {completed}/{processable_total} 条'})}\n\n"
      except Exception as outer_err:
        logger.error(f"[相似度对比] 生成器异常: {str(outer_err)}")
        logger.error(traceback.format_exc())
        yield f"data: {json.dumps({'error': f'生成器异常: {str(outer_err)}'})}\n\n"

    return Response(stream_with_context(generate()), mimetype='text/event-stream')


@app.route('/api/export-similarity/<int:task_id>')
def export_similarity(task_id):
    """导出相似度对比结果为 Excel"""
    import io
    from flask import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from urllib.parse import quote

    task = SimilarityTask.query.get_or_404(task_id)
    items = SimilarityItem.query.filter_by(task_id=task_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '相似度对比结果'

    headers = ['ID', '问题', '答案1', '答案2', '相似度', '等级', '关键差异点']
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='344955', end_color='344955', fill_type='solid')
    thin_border = Border(bottom=Side(style='thin', color='CCCCCC'))
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for idx, item in enumerate(items, 1):
        label_display = ''
        if item.similarity_label == 'high':
            label_display = '高'
        elif item.similarity_label == 'medium':
            label_display = '中'
        elif item.similarity_label == 'low':
            label_display = '低'

        ws.append([
            idx, item.question, item.answer1, item.answer2,
            item.similarity_score if item.similarity_score is not None else '',
            label_display,
            item.key_differences or '',
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    col_widths = [6, 40, 40, 40, 10, 8, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = os.path.splitext(task.original_name)[0] + '_相似度对比结果.xlsx'
    encoded = quote(filename)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded}"}
    )


# ==================== 豆包批量查询 API ====================

@app.route('/doubao')
def doubao_page():
    tasks = DoubaoTask.query.order_by(DoubaoTask.created_at.desc()).all()
    auth_status = doubao_auth.check_status()
    return render_template('doubao.html', tasks=tasks, logged_in=auth_status['logged_in'], auth_state=auth_status['state'])


@app.route('/doubao/<int:task_id>')
def doubao_detail(task_id):
    task = DoubaoTask.query.get_or_404(task_id)
    items = DoubaoItem.query.filter_by(task_id=task_id).all()
    auth_status = doubao_auth.check_status()
    return render_template('doubao_detail.html', task=task, items=items, logged_in=auth_status['logged_in'], auth_state=auth_status['state'])


@app.route('/api/upload-doubao', methods=['POST'])
def upload_doubao():
    """上传豆包批量查询文件"""
    if 'file' not in request.files:
        return jsonify({'error': '没有选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400

    if not allowed_file(file.filename, ALLOWED_EXTENSIONS):
        return jsonify({'error': f'不支持的文件格式，仅支持: {", ".join(ALLOWED_EXTENSIONS)}'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if not ext:
        return jsonify({'error': '无法识别文件扩展名'}), 400
    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        items = parse_file(filepath, 'question')
        if not items:
            os.remove(filepath)
            return jsonify({'error': '文件解析失败或文件为空，请确保包含 question/q/问题 列'}), 400

        task = DoubaoTask(
            filename=filename,
            original_name=file.filename,
            file_type=ext,
            record_count=len(items),
        )
        db.session.add(task)
        db.session.flush()

        for item_data in items:
            item = DoubaoItem(
                task_id=task.id,
                question=item_data['question'],
            )
            db.session.add(item)

        db.session.commit()
        return jsonify({
            'id': task.id,
            'message': f'上传成功，解析到 {len(items)} 条查询',
            'count': len(items),
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'解析失败: {str(e)}'}), 500


@app.route('/api/doubao-query-stream', methods=['POST'])
def doubao_query_stream():
    """流式执行豆包查询"""
    from flask import Response, stream_with_context

    data = request.get_json(force=True, silent=True) or {}
    task_id = data.get('task_id')
    item_ids = data.get('item_ids', [])

    def generate():
        browser = None
        try:
            task = DoubaoTask.query.get(task_id)
            if not task:
                yield f"data: {json.dumps({'error': '任务不存在'})}\n\n"
                return

            if not doubao_auth.is_logged_in():
                yield f"data: {json.dumps({'error': '请先登录豆包'})}\n\n"
                return

            task.status = 'running'
            db.session.commit()

            if item_ids:
                items = DoubaoItem.query.filter(DoubaoItem.id.in_(item_ids)).all()
            else:
                items = DoubaoItem.query.filter_by(task_id=task_id).all()

            total = len(items)

            # 初始化浏览器（使用登录持久化上下文）
            try:
                from services.doubao_browser import DoubaoBrowser
                browser = DoubaoBrowser(doubao_auth)
            except Exception as e:
                yield f"data: {json.dumps({'error': f'浏览器初始化失败: {str(e)}'})}\n\n"
                task.status = 'pending'
                db.session.commit()
                return

            completed = 0
            for i, item in enumerate(items, 1):
                try:
                    answer = browser.query(item.question, timeout=120)
                    item.answer = answer
                    item.status = 'success'
                    completed += 1
                    task.completed_count = completed
                    db.session.commit()

                    yield f"data: {json.dumps({'status': 'ok', 'item_id': item.id, 'progress': i, 'total': total, 'data': {'answer': answer[:200] + '...' if len(answer) > 200 else answer}})}\n\n"
                except Exception as e:
                    error_msg = str(e)
                    item.status = 'error'
                    item.error_msg = error_msg
                    db.session.commit()
                    # 检测登录过期
                    if '登录' in error_msg or '过期' in error_msg:
                        yield f"data: {json.dumps({'status': 'session_expired', 'error': '豆包登录已过期，请重新登录', 'progress': i, 'total': total})}\n\n"
                        break
                    yield f"data: {json.dumps({'status': 'error', 'item_id': item.id, 'error': error_msg, 'progress': i, 'total': total})}\n\n"

            task.status = 'completed'
            db.session.commit()
            yield f"data: {json.dumps({'status': 'done', 'message': f'查询完成，成功 {completed}/{total} 条'})}\n\n"

        except Exception as outer_err:
            import traceback
            yield f"data: {json.dumps({'error': f'查询异常: {str(outer_err)}', 'traceback': traceback.format_exc()})}\n\n"
        finally:
            if browser:
                try:
                    browser.close()
                except Exception:
                    pass

    return Response(stream_with_context(generate()), mimetype='text/event-stream')


@app.route('/api/export-doubao/<int:task_id>')
def export_doubao(task_id):
    """导出豆包查询结果为 Excel"""
    import io
    from flask import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from urllib.parse import quote

    task = DoubaoTask.query.get_or_404(task_id)
    items = DoubaoItem.query.filter_by(task_id=task_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '豆包查询结果'

    headers = ['ID', '问题', '豆包回答', '状态', '错误信息']
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='344955', end_color='344955', fill_type='solid')
    thin_border = Border(bottom=Side(style='thin', color='CCCCCC'))
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for idx, item in enumerate(items, 1):
        status_display = {'pending': '待查询', 'success': '成功', 'error': '失败'}.get(item.status, item.status)
        ws.append([
            idx, item.question, item.answer or '',
            status_display, item.error_msg or '',
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    col_widths = [6, 40, 60, 10, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = os.path.splitext(task.original_name)[0] + '_豆包查询结果.xlsx'
    encoded = quote(filename)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded}"}
    )


@app.route('/api/doubao/<int:task_id>', methods=['DELETE'])
def delete_doubao(task_id):
    """删除豆包查询任务"""
    task = DoubaoTask.query.get_or_404(task_id)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], task.filename)
    if os.path.exists(filepath):
        os.remove(filepath)
    db.session.delete(task)
    db.session.commit()
    return jsonify({'message': '删除成功'})


@app.route('/api/doubao-login', methods=['POST'])
def doubao_login():
    """启动豆包登录流程"""
    result = doubao_auth.start_login()
    if 'error' in result:
        return jsonify(result), 400
    return jsonify(result)


@app.route('/api/doubao-login-status')
def doubao_login_status():
    """查询豆包登录状态"""
    return jsonify(doubao_auth.check_status())


@app.route('/api/doubao-cancel-login', methods=['POST'])
def doubao_cancel_login():
    """取消豆包登录"""
    return jsonify(doubao_auth.cancel_login())


@app.route('/api/doubao-logout', methods=['POST'])
def doubao_logout():
    """退出豆包登录"""
    result = doubao_auth.logout()
    if 'error' in result:
        return jsonify(result), 400
    return jsonify(result)


if __name__ == '__main__':
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    app.run(debug=True, host='0.0.0.0', port=5000)
